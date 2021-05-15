from openpyxl import load_workbook #importa a biblioteca de comunicação entre o Excel e o Python
pastas = (r'C:\Users\flavi\Desktop\pythontraining\Meuarquivo.xlsx') #caminho da planilha do Excel já existente

import time #importa a biblioteca tempo

clv = 12/100            #relação entre cliques e visualizações (a cada uma visualização, eu tenho x cliques)
cocl = 3/20             #relação entre compartilhamentos e cliques (a cada um clique, eu tenho x compartilhamentos)
vco = 40                #relação entre visualizações e compartilhamentos
visualizam = 0          #inicia em 0
clicam = 0              #inicia em 0
compartilham = 0        #inicia em 0
visualizaçoes = 0       #inicia em 0

investimento = 0 #variável que conterá futuramente o total de dinheiro investido no anúncio (dias totais x dinheiro investido diariamente)

encontrado = 0 #variável que me informará futuramente se um usuário foi encontrado no nosso banco de dados

#defino as strings como vazias

registrado = ''
registradocoluna2 = ''
registradocoluna3 = ''
registradocoluna4 = ''

#defino essas variáveis (que são variáveis globais) como 0

diferencaanos = 0
diferencameses = 0
diferencadias = 0
mesI_int = 0
mesT_int = 0
diaI_int = 0
diaT_int = 0

verificacao = 1 #variável que irá permitir que o meu WHILE faça um ciclo de execuções específico no código futuramente. Inicialmente ela valerá 1.

print ("\nPara registramos/consultarmos o seu anúncio, precisaremos de algumas informações. Responda-as corretamente.")
global nome_cliente
nome_cliente = input ("Qual o seu nome completo? ")
time.sleep(0.5)
global nome_anuncio
nome_anuncio = input ("Qual o nome do seu anúncio? ")
time.sleep(0.5)
global data_inicio
data_inicio = input ("Quando será/foi a data de início de publicação do seu anúncio? [dd mm aaaa] ")
time.sleep(0.5)
global data_termino
data_termino = input ("Quando será/foi a data de término de publicação do seu anúncio? [dd mm aaaa] ")
time.sleep(0.5)
global invest
invest = input ("Quanto você quer/foi investir/investido diariamente? ")
time.sleep(0.5)

global diastotal

def diferenca_anos(): #conversão da string de data [16 05 2021, por exemplo]:
    anos_inicio_str = data_inicio.split() #separa a string da data de início de publicação do anúncio em ['16' '05' '2021', por exemplo]
    anos_termino_str = data_termino.split() #separa a string da data de término da publicação do anúncio em ['16' '05' '2021', por exemplo]
    #pego a string responsável por me informar os anos: ['16'] --> [0]; ['05'] --> [1]; ['2021'] --> [2];
    anoI = (anos_inicio_str[2])
    anoT = (anos_termino_str[2])
    #converto o valor dos anos em uma variável do tipo inteiro
    anoI_int = int(anoI)
    anoT_int = int(anoT)
    global diferencaanos #defino como global para poder usar essa variável em outros lugares, não só nessa função
    diferencaanos = ((anoT_int) - (anoI_int)) #faço a diferença do ano de término e do ano de início da publcação do anúncio

def diferenca_meses(): #conversão da string de data [16 05 2021, por exemplo]:
    meses_inicio_str = data_inicio.split() #separa a string da data de início de publicação do anúncio em ['16' '05' '2021', por exemplo]
    meses_termino_str = data_termino.split() #separa a string da data de término da publicação do anúncio em ['16' '05' '2021', por exemplo]
    #pego a string responsável por me informar os meses: ['16'] --> [0]; ['05'] --> [1]; ['2021'] --> [2];
    mesI = (meses_inicio_str[1])
    mesT = (meses_termino_str[1])
    global mesI_int #defino como global
    #converto o valor dos meses em uma variável do tipo inteiro
    mesI_int = int(mesI)
    global mesT_int #defino como global
    #converto o valor dos anos em uma variável do tipo inteiro
    mesT_int = int(mesT)
    global diferencameses #defino como global
    diferencameses = ((mesT_int) - (mesI_int)) #faço a diferença do mês de término e do mês de início da publcação do anúncio

def diferenca_dias(): #conversão da string de data [16 05 2021, por exemplo]:
    dias_inicio_str = data_inicio.split() #separa a string da data de início de publicação do anúncio em ['16' '05' '2021', por exemplo]
    dias_termino_str = data_termino.split() #separa a string da data de término de publicação do anúncio em ['16' '05' '2021', por exemplo]
    #pego a string responsável por me informar os meses: ['16'] --> [0]; ['05'] --> [1]; ['2021'] --> [2];
    diaI = (dias_inicio_str[0])
    diaT = (dias_termino_str[0])
    global diaI_int #defino como global
    #converto o valor dos dias em uma variável do tipo inteiro
    diaI_int = int(diaI)
    global diaT_int #defino como global
    #converto o valor dos dias em uma variável do tipo inteiro
    diaT_int = int(diaT)
    global diferencadias #defino como global
    diferencadias = ((diaT_int) - (diaI_int)) #faço a diferença do dia de término e do dia de início da publcação do anúncio

def calculo(): #função destinada a fazer os cálculos de visualizações, cliques e compartilhamentos com base no dinheiro total investido no anúncio
    investimento = int (invest)
    investotal = (investimento * diastotal)

    visualizacoes = (30 * investotal) #3600
    clicam = (visualizacoes * clv) #se 30 (ou mais) visualizam, quantos clicam? Eu faço a relação entre cliques e visualizações * as visualizações que eu já tenho
    compartilham = (clicam * cocl) #se x pessoas compartilham, quantos clicam? Eu faço a relação entre compartilhamentos e cliques * os cliques que eu já tenho
    visualizam = ((4 * compartilham) * vco) #se x pessoas compartilham, quantos finalmente irão visualizar? Eu tenho que a cada um compartilhamento, eu gero 3 novos, e que a cada 4 compartilhamentos eu gero 40 visualizações finais

    visualizacoes = (visualizacoes + visualizam) #aqui eu somo a quantidade de visualizações que eu gerei pelo meu dinheiro investido com a quantidade de visualizações geradas por todas aquelas interações (cliques e compartilhamentos)

    global visualizacoestotal
    global compartilhamtotal
    global clicamtotal

    visualizacoestotal = int(visualizacoes) #o número utilizado não conterá casas decimais
    compartilhamtotal = int(compartilham) #o número utilizado não conterá casas decimais
    clicamtotal = int(clicam) #o número utilizado não conterá casas decimais

def escreveexcel(row1, col1, row2, col2, row3, col3, row4, col4, row5, col5, row6, col6, row7, col7, row8, col8): #eu digito as linhas e colunas de cada célula que eu quero preencher ao chamar a função

    planilha.cell(row = row1, column = col1, value = (nome_cliente)) #essa célula conterá o nome do cliente
    planilha.cell(row = row2, column = col2, value = (nome_anuncio)) #essa célula conterá o nome do anúncio
    planilha.cell(row = row3, column = col3, value = (data_inicio)) #essa célula conterá a data de início da publicação do anúncio
    planilha.cell(row = row4, column = col4, value = (data_termino)) #essa célula conterá a data de término da publicação do anúncio
    planilha.cell(row = row5, column = col5, value = (invest)) #essa célula conterá o investimento (por dia) no anúncio
    planilha.cell(row = row6, column = col6, value = (visualizacoestotal)) #essa célula conterá as visualizações no total
    planilha.cell(row = row7, column = col7, value = (clicamtotal)) #essa célula conterá os cliques no total
    planilha.cell(row = row8, column = col8, value = (compartilhamtotal)) #essa célula conterá os compartilhamentos no total

def registradofuncao():

    #global é uma forma de eu definir que as variáveis de uma função também serão usadas em outras partes do código, não somente na função
    global registrado
    global registradocoluna2
    global registradocoluna3
    global registradocoluna4
    global registradocoluna5
    global registradocoluna7
    global registradocoluna8
    global registradocoluna9
    global encontrado

    for i in range (1, planilha.max_row +1): #faz a leitura das linhas começando pela linha 1
        #planilha.max_row é o número de linhas que eu tenho na planilha do EXCEL
        if ((planilha.cell(row = i, column = 1).value) == (nome_cliente)): #se achar o nome nas linhas da coluna 1, ele provavelmente já está registrado
            encontrado = 1 #defino encontrado igual a 1
            print("Achou em", ((planilha.cell(row = i, column = 1)))) #ao achar, diga aonde o achou
            #quando eu acho um usuário registrado, devo imprimir na tela de execução do programa o restante das suas informações, como nome do anúncio, datas, e etc...
            registrado = ((planilha.cell(row = i, column = 1))) #aqui eu estou dizendo que a linha em que foi achado o usuário já registrado, será mostrada coluna a coluna:
            registradocoluna2 = ((planilha.cell(row = i, column = 2))) #coluna 2, linha do usuário encontrado
            registradocoluna3 = ((planilha.cell(row = i, column = 3))) #coluna 3, linha do usuário encontrado
            registradocoluna4 = ((planilha.cell(row = i, column = 4))) #coluna 4, linha do usuário encontrado
            registradocoluna5 = ((planilha.cell(row = i, column = 5))) #coluna 5, linha do usuário encontrado
            registradocoluna7 = ((planilha.cell(row = i, column = 7))) #coluna 7, linha do usuário encontrado
            registradocoluna8 = ((planilha.cell(row = i, column = 8))) #coluna 7, linha do usuário encontrado
            registradocoluna9 = ((planilha.cell(row = i, column = 9))) #coluna 8, linha do usuário encontrado
            #OBS: eu pulei a coluna 6 porque ela está toda em amarelo e não é lá que eu quero registrar minhas informações, a coluna 6 serviu como design apenas
            time.sleep(1) #aguardo 1 segundo
            #mostro na tela de execução do programa os meus dados do EXCEL de um usuário já registrado
            print("\n""---------------------------------------------------------------------------------------------------------------------------------------------")
            print("\n","                                               ", "DADOS DO ANÚNCIO REGISTRADO")
            print("\n", planilha["A1"].value, " ",planilha["B1"].value, " ", planilha["C1"].value, " ", planilha["D1"].value, " ", planilha["E1"].value)
            print("\n", "    ", registrado.value, "             ", registradocoluna2.value, "                     ", registradocoluna3.value, "                              ", registradocoluna4.value, "                  ", "R$", registradocoluna5.value,)
            print("\n", planilha["G1"].value, " ",planilha["H1"].value, " ", planilha["I1"].value)
            print("\n", "             ", registradocoluna7.value, "                             ", registradocoluna8.value, "                                 ", registradocoluna9.value)
            print("\n""---------------------------------------------------------------------------------------------------------------------------------------------", "\n")
            break

        else:
            print("Não achou")

while(verificacao == 1): #apenas um loop pra ele ter uma noção de quando começar executar esse código

    diferenca_anos() #chamo a função de conversão da string digitada pelo usuário em um valor em anos
    if (diferencaanos > 0):
        diastotal = (365 * diferencaanos) #calculo isso em dias
    elif (diferencaanos == 0):
        diastotal = 0

    diferenca_meses() #chamo a função de conversão da string digitada pelo usuário em um valor em meses
    if (diferencameses > 0):
        diastotal = ((diastotal) + (30 * diferencameses)) #calculo isso em dias + os dias acumulados da função anterior
    elif (diferencameses == 0):
        diastotal = ((diastotal) + 0) #se não houver diferença dos meses, não haverá conversão em dias
    elif (diferencameses < 0): #se a diferença dos meses for negativa, isso significa que um ano passou, e mais dias deverão ser contabilizados
        diastotal = (diastotal + (30 * ((12 - mesI_int) + mesT_int)))

    diferenca_dias()
    if (diferencadias > 0):
        diastotal = (diastotal + diferencadias) #somo os dias acumulados com a diferença dos dias registrados pelo usuário
    elif (diferencadias == 0):
        diastotal = ((diastotal) + 0) #se não houver diferença de dias, então, tudo se iguala a 0
    elif (diferencadias < 0):
        diastotal = ((diastotal) + ((30 - diaI_int) + diaT_int)) #mesma lógica dos meses negativos

    calculo() #chamo a função de cálculo de visualizações, compartilhamentos e cliques

    wb = load_workbook(pastas) #abre a planilha do excel já existente
    planilha = wb.worksheets[0] #para utilizar a planilha padrão fornecida pelo objeto Workbook

    registradofuncao() #abro a função de verificação de usuário para conferir se ele já não está registrado
    time.sleep(2) #aguardo 2 segundos

    if ((encontrado == 0) and (planilha['A2'].value == None)): #serve para verificar se há valores na célula A2, se não houver...
        escreveexcel(2,1,2,2,2,3,2,4,2,5,2,7,2,8,2,9) #preenche A2,B2,C2,D2,E2,G2,H2,I2
        wb.save(pastas)  #salva a planilha do Excel depois de todas as alterações feitas pelo código

    elif ((encontrado == 0) and (planilha['A2'].value != None)): #se houver algo na célula A2...
        for linha in range (3, planilha.max_row, +1): #começa a procurar por células vazias
            if ((planilha.cell(row = linha, column = 1).value) == (None)): #se encontrar...
                print("Desocupada em", (planilha.cell(row = linha, column = 1))) #informa que está vazia e começa o processo de registro de usuário
                print("Registrando...")
                pergunta = input("Você quer ser registrado no nosso banco de dados? ")

                if (pergunta == "Sim"):
                    escreveexcel(linha, 1, linha, 2, linha, 3, linha, 4, linha, 5, linha, 7, linha, 8, linha, 9) #preenche
                    time.sleep(1)
                    wb.save(pastas)
                    print("Registrado!")
                    break #depois que o usuário deu a sua resposta e foi encontrada uma célula vazia para colocar as suas informações, o break sai do laço de repetição e para de procurar células vazias

                elif (pergunta == "Não"):
                    time.sleep(1)
                    wb.save(pastas)
                    print("Não Registrado!")
                    break #depois que o usuário deu a sua resposta e foi encontrada uma célula vazia para colocar as suas informações, o break sai do laço de repetição e para de procurar células vazias

            else:
                print("Linha ocupada") #enquanto não encontra uma célula vazia, ele escreve isso... até encontrar alguma

    time.sleep(0.5) #espera meio segundo
    registrado = '' #retorna a variável em sua condição inicial
    encontrado = 0 #retorna a variável em sua condição inicial
    verificacao = 2 #encerra o WHILE (e inicia um ciclo novo toda vez que o programa é executado)
